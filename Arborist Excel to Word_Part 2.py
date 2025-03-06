# Head to my github 

from docxcompose.composer import Composer
from docx import Document as Document_compose
import os
import os.path
from docx import Document
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, Fill
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook
import shutil
import pandas as pd
import sys
from docx.shared import RGBColor
from spire.doc.common import *
import docx 

#######################
# Determine variables #
#######################
wb = load_workbook("Tree.xlsx")
sheet = wb['Sheet1']
row_count = sheet.max_row - 1
cwd = os.getcwd()
combinedwordfoldername = "docs"
docpath = cwd + "/" + combinedwordfoldername
filename_master = "combined.docx"     # filename_master is name of the file you want to merge the docx file into #

###############################
# Creation of filename_master #
###############################
shutil.copy(cwd + "/Tree1.docx", cwd + "/" + filename_master)

#####################################################################
# Creation of seperate folder to store all TreeXX.docx word documents #
#####################################################################
try:
    os.makedirs(docpath)
except FileExistsError:
    print(docpath + " already exists")
    print("Please either change " + combinedwordfoldername + " or delete the mentioned folder\nThank you!")
    exit()

##################################
# Move .docx files to new folder #
##################################
import os
import shutil

for i in range(row_count):
    os.rename(cwd + "/Tree"+ str(i+1) +".docx", 
        docpath + "/Tree" + str(i+1) + ".docx")
    print("Tree" + str(i+1) + ".docx" + "moved from working directory to folder " + combinedwordfoldername)


###################################################
# Appending word documents together into one file #
###################################################
master = Document_compose(filename_master)
composer = Composer(master)

for ravana in range(row_count):
    #combined = "combined.docx"
    doc2 = Document_compose(docpath +
        "/"+ 
        "Tree" +
        str(ravana+1) +
        ".docx")
    #append the doc2 into the master using composer.append function
    composer.append(doc2)
    print("Tree" + str(ravana+1) + ".docx" + " has been added to" + filename_master "\n")

#Save the combined docx with a name
composer.save(filename_master)
print("Combined document has been saved under the file name: " + filename_master)


